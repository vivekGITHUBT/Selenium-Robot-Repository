import openpyxl
wk=openpyxl.load_workbook("E:\\TestData.xlsx")
def Fetch_Number_Of_Rows(Sheetname):
    sh=wk[Sheetname]
    return sh.max_row
def fetch_cell_Data(Sheetname,row,cell):
    sh=wk[Sheetname]
    Cell=sh.cell(int(row),int(cell))
    return Cell.value
##print(Fetch_Number_Of_Rows("Sheet1"))
##print(fetch_cell_Data("Sheet1",1,1))